import openpyxl
import datetime
import os


def deposit():

    date = str(datetime.date.today())

    amount = input("Deposit:  ")
    while(not amount.isnumeric()):
        amount = input("Money? ")
    amount = int(amount)

    withdraw = input("Withdraw: ")
    while(not withdraw.isnumeric()):
        withdraw = input("Money? ")
    withdraw = int(withdraw)

    balance = amount-withdraw

    comments = ""  # default
    comments = input("Any Comments? ")

    return [date, amount, withdraw, balance, comments]


def main():

    path = "spendLogs.xlsx"
    dirname = os.path.dirname(__file__)
    paths = os.path.join(dirname, path)

    # if the file already not present, create it
    if (not os.path.exists(paths)):
        file = openpyxl.Workbook()  # creating a new workbook
        sheet = file.active
        sheet.title = "Log1"

        # writing the headers
        headers = ["Date", "Deposit", "Withdraw", "Balance", "Comments"]
        for cols in range(0, len(headers)):
            temp = sheet.cell(row=1, column=cols+1)
            temp.value = headers[cols]

        print("Your Log to Track is Ready! Let's get started! \n")
        file.save(paths)

    else:
        print("Welcome Back! Let's make new saving!...\n")

    response = 1

    while (response != 0):

        # opening the excel file
        logs = openpyxl.load_workbook(paths)
        logsheet = logs.active

        # total number of rows
        totalEntries = logsheet.max_row

        print("\nEnter:")
        temp = "\n----1 to Have a look at your monthly saving : " if totalEntries > 1 else ""
        print(temp)
        print("----2 to add today's record : ")
        print("----3 to modify a record : ")
        print("----4 to fetch a record : ")

        print("----0 to exit\n")

        response = int(input("Option: "))

        if (response not in [1, 2, 3, 4, 0]) or (response == 1 and totalEntries == 1):
            os.system('cls' if os.name == 'nt' else 'clear')
            print("\nWhoops! Not a valid option...\n")
        else:
            if response == 2:

                newDate = str(datetime.date.today())
                flag = False
                for i in range(1, totalEntries):
                    temp = logsheet.cell(row=i, column=1)
                    if newDate == temp.value:
                        flag = True
                        i = i
                        break
                if flag == False:

                    newData = deposit()
                    for cols in range(0, len(newData)):
                        temp = logsheet.cell(row=totalEntries+1, column=cols+1)
                        temp.value = newData[cols]

                    os.system('cls' if os.name == 'nt' else 'clear')
                else:
                    print(
                        "\nWhoops! The Entry already Exists..\nModify it if you want..")

                logs.save(paths)

            elif response == 3:
                newDate = input("Which Date? YYYY-MM-DD: ")
                while(True):
                    temp = newDate.count("-")
                    if not(temp == 2 and len(newDate) == 10):
                        newDate = input("Format YYYY-MM-DD: ")
                    else:
                        break

                flag = False
                for i in range(1, totalEntries):
                    temp = logsheet.cell(row=i, column=1)
                    if newDate == temp.value:
                        flag = True
                        i = i
                        break
                if flag == False:
                    print("\nWhoops! No such Entry!... ")
                else:
                    modified = deposit()
                    modified[0] = newDate
                    for index in range(0, len(modified)):
                        temp = logsheet.cell(row=i, column=index+1)
                        temp.value = modified[index]

                    os.system('cls' if os.name == 'nt' else 'clear')
                    print("Successfully Modified!\n")

                logs.save(paths)

            elif response == 1:

                month = input("Which Month then? MM ")
                while not(month.isnumeric()) or not(1 <= int(month) <= 12):
                    month = input("Enter month Number: ")

                balanceList = []
                for row in range(2, totalEntries):
                    temp = logsheet.cell(row=row, column=1)
                    mnth = temp.value.split("-")
                    mnth = mnth[1]
                    if mnth == month:
                        temp = logsheet.cell(row=row, column=4)
                        balanceList.append(temp.value)

                 os.system('cls' if os.name == 'nt' else 'clear')
                print("\nThe Balance you saved : "+str(sum(balanceList)))
                

            elif response == 4:
                newDate = input("Which Date? YYYY-MM-DD: ")
                while(True):
                    temp = newDate.count("-")
                    if not(temp == 2 and len(newDate) == 10):
                        newDate = input("Format YYYY-MM-DD: ")
                    else:
                        break

                flag = False
                for i in range(1, totalEntries):
                    temp = logsheet.cell(row=i, column=1)
                    if newDate == temp.value:
                        flag = True
                        i = i
                        break
                if flag == False:
                    print("\nWhoops! No such Entry!... ")
                else:
                    os.system('cls' if os.name == 'nt' else 'clear')
                    for col in range(1, 5):
                        temp = logsheet.cell(row=i, column=col)
                        print(temp.value, end=" ")



        logs.save(paths)


main()
